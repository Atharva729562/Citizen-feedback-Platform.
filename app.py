from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
from flask_mail import Mail, Message
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__, template_folder="templates")
CORS(app)

# Configure Flask-Mail
app.config['MAIL_SERVER'] = 'smtp.gmail.com'  # Use your email provider's SMTP server
app.config['MAIL_PORT'] = 587  # Typically 587 for TLS
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = "mahajanatharva49@gmail.com"  # Replace with your email
app.config['MAIL_PASSWORD'] = "vuvolwyvigduklbu"  # Use an App Password, NOT your actual password
app.config['MAIL_DEFAULT_SENDER'] = "mahajanatharva49@gmail.com"  # Replace with your email

mail = Mail(app)

# Directory to save Excel files
DATA_DIR = "data"
os.makedirs(DATA_DIR, exist_ok=True)

# File paths for Excel sheets
SURVEY_FILE = os.path.join(DATA_DIR, 'survey_data.xlsx')
FEEDBACK_FILE = os.path.join(DATA_DIR, 'feedback_data.xlsx')

# Function to get rating statistics
def get_rating_stats():
    try:
        workbook = load_workbook(SURVEY_FILE)
        sheet = workbook.active

        healthcare_ratings = []
        public_services_ratings = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            healthcare_ratings.append(row[1])  # Column index for Healthcare
            public_services_ratings.append(row[3])  # Column index for Public Services

        # Count occurrences of each rating
        def calculate_percentage(ratings):
            total_responses = len(ratings)
            rating_count = {str(i): ratings.count(str(i)) for i in range(1, 11)}
            if total_responses > 0:
                return {k: f"{(v / total_responses) * 100:.1f}%" for k, v in rating_count.items()}
            return {str(i): "0%" for i in range(1, 11)}

        healthcare_percentages = calculate_percentage(healthcare_ratings)
        public_services_percentages = calculate_percentage(public_services_ratings)

        return {
            "healthcare": healthcare_percentages,
            "public_services": public_services_percentages
        }
    except Exception as e:
        return {"error": str(e)}

# Initialize Excel files if they don't exist
def initialize_excel(file_path, headers):
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        workbook.save(file_path)

# Initialize survey and feedback files
initialize_excel(SURVEY_FILE, ["Question1", "Healthcare", "Education", "Public Services", "Priority", "Rating"])
initialize_excel(FEEDBACK_FILE, ["Name", "Contact", "Email", "Comment"])

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/get-rating-stats', methods=['GET'])
def get_rating_stats_api():
    return jsonify(get_rating_stats())

# Save Survey Data
@app.route('/save-survey', methods=['POST'])
def save_survey():
    survey_data = request.json
    try:
        workbook = load_workbook(SURVEY_FILE)
        sheet = workbook.active
        sheet.append([
            survey_data.get("question1"),
            survey_data.get("healthcare"),
            survey_data.get("education"),
            survey_data.get("publicServices"),
            survey_data.get("priority"),
            survey_data.get("rating")
        ])
        workbook.save(SURVEY_FILE)
        return jsonify({"message": "Survey data saved successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Save Feedback Data
@app.route('/save-feedback', methods=['POST'])
def save_feedback():
    feedback_data = request.json
    user_email = feedback_data.get("email")
    try:
        workbook = load_workbook(FEEDBACK_FILE)
        sheet = workbook.active
        sheet.append([
            feedback_data.get("name"),
            feedback_data.get("contact"),
            user_email,
            feedback_data.get("comment")
        ])
        workbook.save(FEEDBACK_FILE)

          # Send confirmation email
        if user_email:
            send_email(user_email)

        return jsonify({"message": "Feedback data saved successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
 # Function to send email
def send_email(to_email):
    try:
        msg = Message("Citizen Feedback Submission Confirmation",
                      recipients=[to_email])
        msg.body = "Thank you for submitting your feedback. Your response has been recorded successfully!"
        mail.send(msg)
    except Exception as e:
        print(f"Error sending email: {e}")   

if __name__ == '__main__':
    app.run(debug=True, port=5000)


